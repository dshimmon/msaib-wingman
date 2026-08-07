# Tests Excel workbook extraction and routing.

import sys
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SRC_DIRECTORY = PROJECT_ROOT / "src"

sys.path.insert(0, str(SRC_DIRECTORY))

from wingman.core.document_router import extract_document_units
from wingman.core.excel_adapter import extract_excel_units


class ExcelAdapterTests(unittest.TestCase):
    def create_workbook(self, build_workbook):
        temporary_directory = tempfile.TemporaryDirectory()
        self.addCleanup(temporary_directory.cleanup)

        file_path = (
            Path(temporary_directory.name)
            / "test-workbook.xlsx"
        )
        workbook = Workbook()
        build_workbook(workbook)
        workbook.save(file_path)
        workbook.close()

        return file_path

    def create_readable_workbook(self):
        def build_workbook(workbook):
            overview = workbook.active
            overview.title = "Overview"
            overview.append(["Name", None, "Status", None])
            overview.append(["Avery", 42, True, None])
            overview.append([])
            overview.append([])
            overview.append(
                [
                    "Dates",
                    date(2026, 7, 26),
                    datetime(2026, 7, 26, 14, 30),
                ]
            )
            overview.append(
                ["Total", None, "=SUM(B2:B10)"]
            )

            workbook.create_sheet("Blank")

            details = workbook.create_sheet("Details")
            details.append(["Final row", 3.5])

        return self.create_workbook(build_workbook)

    def test_extracts_row_groups_in_worksheet_order(self):
        units = extract_excel_units(
            self.create_readable_workbook()
        )

        self.assertEqual(
            [
                (unit["heading"], unit["location"])
                for unit in units
            ],
            [
                (
                    "Overview",
                    "Sheet Overview, Rows 1-2",
                ),
                (
                    "Overview",
                    "Sheet Overview, Rows 5-6",
                ),
                (
                    "Details",
                    "Sheet Details, Rows 1-1",
                ),
            ],
        )

    def test_preserves_supported_cell_values(self):
        units = extract_excel_units(
            self.create_readable_workbook()
        )

        self.assertEqual(
            units[0]["text"],
            "Name |  | Status\nAvery | 42 | True",
        )
        self.assertEqual(
            units[1]["text"],
            (
                "Dates | 2026-07-26 00:00:00 | "
                "2026-07-26 14:30:00\n"
                "Total |  | =SUM(B2:B10)"
            ),
        )
        self.assertEqual(
            units[2]["text"],
            "Final row | 3.5",
        )

    def test_empty_workbook_raises_clear_error(self):
        def build_workbook(workbook):
            workbook.active.title = "Empty"
            workbook.create_sheet("Also Empty")

        with self.assertRaisesRegex(
            ValueError,
            r"^Excel workbook contains no readable data\.$",
        ):
            extract_excel_units(
                self.create_workbook(build_workbook)
            )

    def test_router_routes_xlsx_files(self):
        file_path = self.create_readable_workbook()

        self.assertEqual(
            extract_document_units(file_path),
            extract_excel_units(file_path),
        )

    def test_router_preserves_unsupported_format_error(self):
        with self.assertRaisesRegex(
            ValueError,
            r"Unsupported document type: \.rtf",
        ):
                extract_document_units("notes.rtf")


if __name__ == "__main__":
    unittest.main()
