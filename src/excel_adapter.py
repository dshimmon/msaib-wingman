# Extracts normalized document units from Excel workbooks.

from datetime import date, datetime, time

from openpyxl import load_workbook


def format_cell_value(value):
    """
    Convert a worksheet cell value to stable readable text.
    """
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.isoformat(
            sep=" ",
            timespec="seconds",
        )

    if isinstance(value, (date, time)):
        return value.isoformat()

    if isinstance(value, str):
        return value.strip()

    return str(value)


def flatten_worksheet_row(row):
    """
    Flatten a row while retaining meaningful internal empty cells.
    """
    cell_values = [
        format_cell_value(cell.value)
        for cell in row
    ]

    while cell_values and not cell_values[-1]:
        cell_values.pop()

    if not cell_values:
        return None

    return " | ".join(cell_values)


def extract_excel_units(file_path):
    """
    Extract blank-row-delimited units from an .xlsx workbook.
    """
    workbook = load_workbook(
        file_path,
        data_only=False,
        read_only=True,
    )
    units = []

    try:
        for worksheet in workbook.worksheets:
            current_rows = []
            start_row = None
            end_row = None

            def append_current_group():
                if not current_rows:
                    return

                units.append(
                    {
                        "heading": worksheet.title,
                        "location": (
                            f"Sheet {worksheet.title}, "
                            f"Rows {start_row}-{end_row}"
                        ),
                        "text": "\n".join(current_rows),
                    }
                )

            for row_number, row in enumerate(
                worksheet.iter_rows(),
                start=1,
            ):
                row_text = flatten_worksheet_row(row)

                if row_text is None:
                    append_current_group()
                    current_rows = []
                    start_row = None
                    end_row = None
                    continue

                if start_row is None:
                    start_row = row_number

                current_rows.append(row_text)
                end_row = row_number

            append_current_group()
    finally:
        workbook.close()

    if not units:
        raise ValueError(
            "Excel workbook contains no readable data."
        )

    return units
